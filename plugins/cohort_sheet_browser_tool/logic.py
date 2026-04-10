from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

from app.models import SpreadsheetSourceResponse
from app.services.tool_runner import discover_tools

MAX_GRID_ROWS = 100


def _load_workbook_tables(spreadsheet_path: str) -> tuple[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required for spreadsheet intake. Create the conda environment from environment.yml before using spreadsheet review."
        ) from exc

    workbook = load_workbook(filename=spreadsheet_path, read_only=True, data_only=True)
    suffix = Path(spreadsheet_path).suffix.lower().lstrip(".") or "xlsx"
    sheet_tables: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
          first_row = next(rows_iter)
        except StopIteration:
          continue
        if first_row is None:
          continue
        columns = [str(value).strip() if value is not None else f"column_{index + 1}" for index, value in enumerate(first_row)]
        if not any(columns):
          columns = [f"column_{index + 1}" for index, _ in enumerate(first_row)]
        normalized_columns: list[str] = []
        seen: dict[str, int] = {}
        for index, column in enumerate(columns):
            base = column or f"column_{index + 1}"
            count = seen.get(base, 0)
            seen[base] = count + 1
            normalized_columns.append(base if count == 0 else f"{base}_{count + 1}")
        data_rows: list[dict[str, str]] = []
        for row in rows_iter:
            payload = {
                normalized_columns[index]: ("" if value is None else str(value))
                for index, value in enumerate(row[: len(normalized_columns)])
            }
            if any(str(value or "").strip() for value in payload.values()):
                data_rows.append(payload)
        if data_rows:
            sheet_tables.append(
                {
                    "sheet_name": worksheet.title,
                    "columns": normalized_columns,
                    "rows": data_rows,
                }
            )
    workbook.close()
    return suffix, sheet_tables


def _normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(name).lower()).strip("_")


def _name_matches(name: str, patterns: tuple[str, ...]) -> bool:
    normalized = _normalize_name(name)
    return any(pattern in normalized for pattern in patterns)


def _is_float_like(value: str) -> bool:
    try:
        float(value)
        return True
    except Exception:
        return False


def _infer_type(values: list[str]) -> str:
    non_empty = [value for value in values if str(value or "").strip()]
    if not non_empty:
        return "categorical"
    float_like = sum(1 for value in non_empty if _is_float_like(str(value)))
    date_like = sum(1 for value in non_empty if re.match(r"^\d{4}-\d{1,2}-\d{1,2}", str(value)))
    if float_like == len(non_empty):
        if all(float(str(value)).is_integer() for value in non_empty):
            return "integer"
        return "float"
    if date_like >= max(1, math.ceil(len(non_empty) * 0.7)):
        return "date-like"
    unique_count = len(set(str(value) for value in non_empty))
    if unique_count > max(20, len(non_empty) * 0.6):
        return "free-text"
    return "categorical"


def _build_profiles(columns: list[str], rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    profiles: list[dict[str, Any]] = []
    for column in columns:
        values = [str(row.get(column, "") or "") for row in rows]
        non_empty = [value for value in values if value.strip()]
        missing_count = len(values) - len(non_empty)
        inferred = _infer_type(values)
        profiles.append(
            {
                "name": column,
                "inferred_type": inferred,
                "non_empty_count": len(non_empty),
                "missing_count": missing_count,
                "missing_rate": missing_count / max(len(values), 1),
                "unique_count": len(set(non_empty)),
            }
        )
    return profiles


def _infer_roles(columns: list[str], profiles: list[dict[str, Any]]) -> dict[str, list[str]]:
    role_map = {
        "subject_id_columns": ("subject", "patient_id", "participant", "screening", "mrn", "person_id", "subjectno"),
        "visit_columns": ("visit", "timepoint", "epoch", "cycle"),
        "site_columns": ("site", "center", "hospital"),
        "arm_columns": ("arm", "group", "cohort", "treatment"),
        "date_columns": ("date", "time", "datetime"),
        "outcome_columns": ("outcome", "response", "status", "grade", "severity"),
    }
    by_name = {profile["name"]: profile for profile in profiles}
    roles: dict[str, list[str]] = {key: [] for key in role_map}
    for column in columns:
        inferred = str((by_name.get(column) or {}).get("inferred_type", ""))
        for role, patterns in role_map.items():
            if _name_matches(column, patterns):
                if role == "date_columns" and inferred not in {"date-like", "integer", "float", "categorical"}:
                    continue
                roles[role].append(column)
    return roles


def _pick_subject_column(rows: list[dict[str, str]], roles: dict[str, list[str]], columns: list[str]) -> str | None:
    if roles["subject_id_columns"]:
        return roles["subject_id_columns"][0]
    preferred = ("subject", "patient", "participant", "person", "mrn", "screen", "id", "subjectno")
    for column in columns:
        normalized = _normalize_name(column)
        if any(token in normalized for token in preferred):
            return column
    return None


def _classify(file_name: str, rows: list[dict[str, str]], columns: list[str], profiles: list[dict[str, Any]], roles: dict[str, list[str]], suffix: str) -> dict[str, Any]:
    row_count = len(rows)
    cohort_score = 0
    single_score = 0
    rationale: list[str] = []
    subject_column = _pick_subject_column(rows, roles, columns)
    subject_unique = 0
    if subject_column:
        subject_values = {
            str(row.get(subject_column, "") or "").strip()
            for row in rows
            if str(row.get(subject_column, "") or "").strip()
        }
        subject_unique = len(subject_values)
        if subject_unique > 1:
            cohort_score += 5
            rationale.append(f"`{subject_column}` has {subject_unique} unique patient/subject identifiers.")
        elif subject_unique == 1:
            single_score += 3
            rationale.append(f"`{subject_column}` carries a single patient/subject identifier.")
    if row_count > 20:
        cohort_score += 3
        rationale.append(f"Table has {row_count} rows, which is more consistent with cohort-style eCRF data.")
    elif row_count <= 3:
        single_score += 3
        rationale.append(f"Table has only {row_count} row(s), which is more consistent with a single-patient worksheet.")
    if any(roles[key] for key in ("visit_columns", "site_columns", "arm_columns")):
        cohort_score += 3
        present = [label for label in ("visit_columns", "site_columns", "arm_columns") if roles[label]]
        rationale.append(f"Detected cohort-style organizational columns: {', '.join(present)}.")
    if suffix in {"xlsx", "xlsm"}:
        cohort_score += 1
        rationale.append("Excel workbook intake slightly favors cohort/eCRF interpretation.")
    analysis_mode = "ambiguous"
    if cohort_score >= single_score + 2:
        analysis_mode = "cohort"
    elif single_score >= cohort_score + 2:
        analysis_mode = "single-patient"
    return {
        "analysis_mode": analysis_mode,
        "cohort_score": cohort_score,
        "single_patient_score": single_score,
        "subject_column": subject_column,
        "subject_unique_count": subject_unique,
        "visit_columns": roles["visit_columns"],
        "site_columns": roles["site_columns"],
        "arm_columns": roles["arm_columns"],
        "rationale": rationale,
        "row_count": len(rows),
        "column_count": len(columns),
    }


def _missingness_summary(profiles: list[dict[str, Any]]) -> dict[str, Any]:
    ranked = sorted(profiles, key=lambda item: item["missing_rate"], reverse=True)
    return {
        "top_missing_columns": [
            {
                "column": item["name"],
                "missing_rate": item["missing_rate"],
                "missing_count": item["missing_count"],
                "non_empty_count": item["non_empty_count"],
            }
            for item in ranked[:8]
        ]
    }


def _value_counts(rows: list[dict[str, str]], column: str | None, limit: int = 10) -> list[dict[str, Any]]:
    if not column:
        return []
    counts: dict[str, int] = {}
    for row in rows:
        value = str(row.get(column, "") or "").strip() or "(missing)"
        counts[value] = counts.get(value, 0) + 1
    return [{"label": label, "count": count} for label, count in sorted(counts.items(), key=lambda item: item[1], reverse=True)[:limit]]


def _cohort_summary(rows: list[dict[str, str]], profiles: list[dict[str, Any]]) -> dict[str, Any]:
    categorical = [item for item in profiles if item["inferred_type"] == "categorical" and item["unique_count"] > 0]
    numeric = [item for item in profiles if item["inferred_type"] in {"integer", "float"}]
    categorical_breakdowns = []
    for profile in sorted(categorical, key=lambda item: item["unique_count"])[:3]:
        counts = _value_counts(rows, profile["name"], limit=5)
        categorical_breakdowns.append({"column": profile["name"], "top_values": counts})
    numeric_breakdowns = [{"column": item["name"], "summary": {"missing_count": item["missing_count"], "unique_count": item["unique_count"]}} for item in numeric[:3]]
    return {
        "record_count": len(rows),
        "field_count": len(profiles),
        "categorical_breakdowns": categorical_breakdowns,
        "numeric_breakdowns": numeric_breakdowns,
    }


def _build_subject_preview(rows: list[dict[str, str]], roles: dict[str, list[str]], columns: list[str]) -> list[dict[str, Any]]:
    subject_column = _pick_subject_column(rows, roles, columns)
    if not subject_column:
        return []
    visit_column = roles["visit_columns"][0] if roles["visit_columns"] else None
    site_column = roles["site_columns"][0] if roles["site_columns"] else None
    arm_column = roles["arm_columns"][0] if roles["arm_columns"] else None
    outcome_column = roles["outcome_columns"][0] if roles["outcome_columns"] else None
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        subject = str(row.get(subject_column, "") or "").strip() or "(missing)"
        grouped.setdefault(subject, []).append(row)
    preview = []
    for subject, subject_rows in list(grouped.items())[:12]:
        latest_row = subject_rows[-1]
        visits = []
        if visit_column:
            visits = sorted(
                {
                    str(item.get(visit_column, "") or "").strip()
                    for item in subject_rows
                    if str(item.get(visit_column, "") or "").strip()
                }
            )
        preview.append(
            {
                "subject_id": subject,
                "record_count": len(subject_rows),
                "site": str(latest_row.get(site_column, "") or "n/a") if site_column else "n/a",
                "arm": str(latest_row.get(arm_column, "") or "n/a") if arm_column else "n/a",
                "latest_outcome": str(latest_row.get(outcome_column, "") or "n/a") if outcome_column else "n/a",
                "visits": visits,
            }
        )
    return preview


def _sheet_domain_name(sheet_name: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", sheet_name.lower()).strip("_")
    return normalized or "sheet"


def _build_artifact(
    rows: list[dict[str, str]],
    columns: list[str],
    profiles: list[dict[str, Any]],
    roles: dict[str, list[str]],
    intake: dict[str, Any],
    cohort: dict[str, Any],
    missingness: dict[str, Any],
    sheet_details: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "overview": {
            "row_count": len(rows),
            "column_count": len(columns),
            "subject_count": intake.get("subject_unique_count", 0),
            "analysis_mode": intake.get("analysis_mode", "ambiguous"),
            "selected_sheet": ((intake.get("table_meta") or {}).get("selected_sheet") or "n/a"),
        },
        "intake": intake,
        "composition": cohort,
        "domains": [
            {
                "sheet_name": item.get("sheet_name", "Sheet"),
                "domain": item.get("domain", _sheet_domain_name(str(item.get("sheet_name", "Sheet")))),
                "row_count": item.get("row_count", 0),
                "subject_count": item.get("subject_count", 0),
                "subject_column": item.get("subject_column", "n/a"),
                "visit_columns": item.get("visit_columns", []),
                "date_columns": item.get("date_columns", []),
            }
            for item in sheet_details
        ],
        "subjects": _build_subject_preview(rows, roles, columns),
        "grid": {"columns": columns, "rows": rows[:MAX_GRID_ROWS], "row_count": len(rows)},
        "schema_highlights": [
            {
                "name": item["name"],
                "inferred_type": item["inferred_type"],
                "missing_count": item["missing_count"],
                "unique_count": item["unique_count"],
            }
            for item in profiles[:8]
        ],
        "roles": roles,
        "missingness": missingness,
    }


def analyze_spreadsheet_source(spreadsheet_path: str, file_name: str | None = None) -> SpreadsheetSourceResponse:
    path = Path(spreadsheet_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Spreadsheet source not found: {path}")
    suffix, sheet_tables = _load_workbook_tables(str(path))
    if not sheet_tables:
        raise ValueError("The workbook does not contain any non-empty sheets.")

    sheet_details: list[dict[str, Any]] = []
    studio_cards: list[dict[str, Any]] = []
    artifacts: dict[str, Any] = {}
    summary_lines = [
        f"This workbook contains **{len(sheet_tables)}** non-empty sheet(s), each analyzed as an individual cohort view.",
        "",
        "## Sheet overview",
        "",
    ]

    for sheet_table in sheet_tables:
        sheet_rows = list(sheet_table.get("rows") or [])
        sheet_columns = list(sheet_table.get("columns") or [])
        sheet_name = str(sheet_table.get("sheet_name") or "Sheet")
        profiles = _build_profiles(sheet_columns, sheet_rows)
        roles = _infer_roles(sheet_columns, profiles)
        intake = _classify(path.name, sheet_rows, sheet_columns, profiles, roles, suffix)
        cohort = _cohort_summary(sheet_rows, profiles)
        missingness = _missingness_summary(profiles)
        detail = {
            "sheet_name": sheet_name,
            "domain": _sheet_domain_name(sheet_name),
            "row_count": len(sheet_rows),
            "subject_count": intake.get("subject_unique_count", 0),
            "subject_column": intake.get("subject_column", "n/a"),
            "visit_columns": roles.get("visit_columns", []),
            "date_columns": roles.get("date_columns", []),
        }
        sheet_details.append(detail)
        card_id = f"sheet::{sheet_name}::cohort_browser"
        sheet_table_meta = {
            "workbook_format": suffix,
            "sheet_names": [item.get("sheet_name") for item in sheet_tables],
            "selected_sheet": sheet_name,
            "sheet_name": sheet_name,
            "sheet_details": sheet_details,
        }
        studio_cards.append({"id": card_id, "title": sheet_name, "subtitle": "Cohort Browser", "base_id": "cohort_browser"})
        artifacts[card_id] = _build_artifact(
            sheet_rows,
            sheet_columns,
            profiles,
            roles,
            {**intake, "table_meta": sheet_table_meta},
            cohort,
            missingness,
            sheet_details,
        )
        summary_lines.append(
            f"- **{sheet_name}**: {len(sheet_rows)} row(s), {len(sheet_columns)} column(s), mode `{intake['analysis_mode']}`, subjects {intake.get('subject_unique_count', 'n/a')}."
        )

    selected_sheet_name = str(sheet_tables[0]["sheet_name"])
    draft_answer = "\n".join(
        summary_lines
        + [
            "",
            f"**Selected sheet:** `{selected_sheet_name}`",
            "",
            "Open the Studio card to browse a sheet-level cohort view, or ask about a specific sheet by name.",
        ]
    )
    return SpreadsheetSourceResponse(
        analysis_id="",
        source_spreadsheet_path=str(path),
        file_name=file_name or path.name,
        workbook_format=suffix,
        sheet_names=[str(item.get("sheet_name") or "Sheet") for item in sheet_tables],
        selected_sheet=selected_sheet_name,
        sheet_count=len(sheet_tables),
        sheet_details=sheet_details,
        studio_cards=studio_cards,
        artifacts=artifacts,
        warnings=[],
        draft_answer=draft_answer,
        used_tools=["cohort_sheet_browser_tool"],
        tool_registry=discover_tools(),
    )


def execute(payload: dict[str, object]) -> dict[str, object]:
    spreadsheet_path = str(payload.get("spreadsheet_path") or "").strip()
    if not spreadsheet_path:
        raise ValueError("`spreadsheet_path` is required.")
    file_name = str(payload.get("file_name") or Path(spreadsheet_path).name).strip()
    analysis = analyze_spreadsheet_source(spreadsheet_path, file_name=file_name)
    return {"analysis": analysis.model_dump()}
